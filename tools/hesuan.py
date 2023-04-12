# -*- coding:utf-8 -*-
# ！/usr/bin/env python
# @Time : 2022/9/5 10:54
# @Author : waxberry
# @File : hesuan.py
# @Software : PyCharm

# 录入数据
from datetime import datetime
from turtle import left

my_nucleic = {
    'date': [datetime.strptime(datetime(2022, 8, i+1), '%Y-%m-%d') for i in range(31)],
    'nucleic': [1, 1, 0, 1, 0, 1, 0, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
}

# 使用openpyxl创建表格
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active


# 定义表格初始化和单元格设置的函数
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def init_sheet(ws):
    for r in range(100):
        for c in range(100):
            ws.cess(row=r+1, column=c+1).fill = PatternFill('solid', fgColor='000000')

def set_cell_style(ws, r, c, color):
    ws.cell(row=r, column=c).fill = PatternFill('solid', fgColor=color)
    ws.cell(row=r, column=c).font = Font(name='微软雅黑', size=14, bold=True)
    ws.cell(row=r, column=c).alignment = Alignment(horizontal='right', vertical='center')
    side = Side(style='medium', color='004B3C')
    ws.cell(row=r, column=c).border = Border(top=side, bottom=side, left=side, right=side)

# 实现日历
import calendar
init_sheet(ws)

side = Side(style='medium', color='004B3C')
for col in range(7):
    ws.cell(row=1, column=col+1).border = Border(top=side, bottom=side, left=side, right=side)

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

ws.cell(row=1, column=1).value = '2022年8月'
set_cell_style(ws, r=1, c=1, color='418CFA')

title_data = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']

for col in range(7):
    ws.cell(row=2, column=col+1).value = title_data[col]
    set_cell_style(ws, r=2, c=col+1, color='418CFA')

mouthday = calendar.monthcalendar(2022, 8)

col, row = mouthday[0], 3
for i in range(len(my_nucleic['date'])):
    if col < 7:
        ws.cell(row=row, column=col+1).value = i+1
        col += 1
    else:
        col = 0
        row += 1
        ws.cell(row=row, column=col + 1).value = i+1
        col += 1

    set_cell_style(ws, r=row, c=col, color='000000')

    if my_nucleic['nucleic'][i] == 1:
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor='009B3C')

for i in range(1, row + 1):
    ws.row_dimensions[i].height = 30

wb.save(filename='show_august_nucleic.xlsx')
wb.close()




# 按年
from pyecharts import options as opts
from pyecharts.charts import Calendar
import pandas as pd

nucleic_df = pd.DataFrame()


def made_data(param, param1):
    pass


for i in range(12):
    mouth_nucleic = made_data(2022, i+1)
    mouth_df = pd.DataFrame(mouth_nucleic)
    nucleic_df = pd.concat([nucleic_df, mouth_df])

data = [[row_data['date'], row_data['nucleic']] for row_index, row_data in nucleic_df.iterrows()]
cal = Calendar(init_opts=opts.InitOpts(width='900px', height='500px'))
cal.add('', data, calendar_opts=opts.CalendarOpts(range_='2022',
                                                  daylabel_opts=opts.CalendarDayLabelOpts(first_dat=1, name_map='cn'))
        ).set_series_opts(
    label_opts=opts.LabelOpts(font_size=12)
).set_global_opts(
    title_opts=opts.TitleOpts(title='核酸检测日历', pos_left='450', pos_top='0',
                              title_textstyle_opts=opts.TextStyleOpts(color='black', font_size=16)),
    visualmap_opts=opts.VisualMapOpts(
        max_=1, min_=0, orient='horizontal', is_piecewise=False,
        range_color=['white', 'white', 'green'], pos_top='250px', pos_left='50px'
    ),
).render('my_nucleic.html')
