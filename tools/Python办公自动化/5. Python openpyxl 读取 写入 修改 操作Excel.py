# -*- coding:utf-8 -*-
# ！/usr/bin/env python
# @Time : 2023/7/28 17:17
# @Author : waxberry
# @File : 5. Python openpyxl 读取 写入 修改 操作Excel.py
# @Software : PyCharm


# 5.1 openpyxl 基本操作

# 2.打开文件
# （1）新建
import datetime

from openpyxl import Workbook
# 实例化
wb = Workbook()
# 激活worksheet
ws = wb.active

# （2）打开已有
from openpyxl import load_workbook

wb = load_workbook('name.xlsx')

# 3.写入数据
# 方式一：数据可以直接分配到单元格中(可以输入公式)
ws['A1'] = 42
# 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
ws.append([1, 2, 3])
# 方式三：Python 类型会被自动转换
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")

# 4.创建表（sheet）
# 方式一：插入到最后(default)
ws1 = wb.create_sheet('my_sheet')
# 方式二：插入到最开始的位置
ws2 = wb.create_sheet('my_sheet', 0)

# 5.选择表（sheet）
# sheet 名称可以作为 key 进行索引
ws3 = wb['new title']
ws4 = wb.get_sheet_name('new_title')
ws is ws3 is ws4
True

# 6.查看表名（sheet）
# 显示所有表名
print(wb.sheetnames)
# 遍历所有表
for sheet in wb:
    print(sheet.title)

# 7.访问单元格（cell）
# （1）单个单元格访问
# 方法一
c = ws['A4']
# 方法二：row 行；column 列
d = ws.cell(row=4, column=2, value=10)
# 方法三：只要访问就创建
for i in range(1, 101):
    for j in range(1, 101):
        ws.cell(row=i, column=j)

# （2）多个单元格访问
# 通过切片
cell_range = ws['A1: C2']
# 通过行(列)
colC = ws['C']
col_range = ws['C:D']
rowl0 = ws[10]
row_range = ws[5:10]
# 通过指定范围(行 → 行)
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)

# 通过指定范围(列 → 列)
for column in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in column:
        print(cell)

# 遍历所有 方法一
ws = wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)
# 遍历所有 方法二
tuple(ws.columns)

# 8.保存数据
wb.save('name.xlsx')

# 9.其它
# （1）改变sheet标签按钮颜色
ws.sheet_properties.tabColor = '1072BA'# 色值为RGB16进制值

# （2）获取最大行，最大列
# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)

# （3）获取每一行每一列
'''sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。sheet.columns类似，不过里面是每个tuple是每一列的单元格。'''
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)
# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)

# （4）根据数字得到字母，根据字母得到数字
from openpyxl.utils import get_column_letter, column_index_from_string
# 根据列的数字返回字母
print(get_column_letter(2))# B
# 根据字母返回列的数字
print(column_index_from_string('D'))# 4

# （5）删除工作表
# 方式一
wb.remove(sheet)
# 方式二
del wb[sheet]

# （6）矩阵置换
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10]
]
list(zip(*rows))

# out
[('Number', 2, 3, 4, 5, 6, 7),
('data1', 40, 40, 50, 30, 25, 50),
('data2', 30, 25, 30, 10, 5, 10)]
# 注意 方法会舍弃缺少数据的列(行)
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40      ],    # 这里少一个数据
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
# out
[('Number', 2, 3, 4, 5, 6, 7), ('data1', 40, 40, 50, 30, 25, 50)]

# 10.设置单元格风格
# （1）需要导入的类
from openpyxl.styles import Font, colors, Alignment

# （2）字体
'''下面的代码指定了等线24号，加粗斜体，字体颜色红色。直接使用cell的font属性，将Font对象赋值给它。'''
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
sheet['A1'].font = bold_itatic_24_font

# （3）对齐方式
'''也是直接使用cell的属性aligment，这里指定垂直居中和水平居中。除了center，还可以使用right、left等等参数'''
# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

# （4）设置行高和列宽
# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

# （5）合并和拆分单元格
# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
sheet.merge_cells('A1:C3') # 合并一个矩形区域中的几个单元格
# 拆分单元格的代码。拆分后，值回到A1位置
sheet.unmerge_cells('A1:C3')

# 11.示例代码
import datetime
from random import choice
from time import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 设置文件 mingc
addr = 'openxlsx.xlsx'
# 打开文件
wb = load_workbook(addr)
# 创建一张新表
ws = wb.create_sheet()
# 第一行输入
ws.append(['TIME', 'TITLE', 'A-Z'])

# 输入内容（500行数据）
for i in range(500):
    TIME = datetime.datetime.now().strftime('%H-%M-%S')
    TITLE = str(time())
    A_Z = get_column_letter(choice(range(1, 50)))
    ws.append([TIME, TITLE, A_Z])

# 获取最大行
row_max = ws.max_row
# 获取最大列
column_max = ws.max_column
# 把上面写入内容打印在控制台
for j in ws.rows:# we.rows 获取每一行数据
    for n in j:
        print(n.value, end='\t')# n.value 获取单元格的值
    print()
# 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
wb.save(addr)

# 5.2 openpyxl生成2D图表
# 示例代码：
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference

wb = Workbook(write_only=True)
ws = wb.create_sheet()

rows = [
    ('Number', 'Batch 1', 'Batch 2'),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30),
]

for row in rows:
    ws.append(row)

chart1 = BarChart()
chart1.type = 'col'
chart1.style = 10
chart1.title = 'Bar Chart'
chart1.y_axis.title = 'Test number'
chart1.x_axis.title = 'Sample length(mm)'

data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)
cats = Reference(ws, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
ws.add_chart(chart1, 'A10')

from copy import deepcopy

chart2 = deepcopy(chart1)
chart2.style = 11
chart2.type = 'bar'
chart2.title = 'Horizontal Bar Chart'
ws.add_chart(chart2, 'G10')

chart3 = deepcopy(chart1)
chart3.type = 'col'
chart3.style = 12
chart3.grouping = 'stacked'
chart3.overlap = 100
chart3.title = 'Stacked Chart'
ws.add_chart(chart3, 'A27')

chart4 = deepcopy(chart1)
chart4.type = 'Bar'
chart4.style = 13
chart4.grouping = 'percentStacked'
chart4.overlap = 100
chart4.title = 'Percent Stacked Chart'
ws.add_chart(chart4, 'G27')

wb.save('bar.xlsx')

# 5.3 openpyxl生成3D图表
# 示例代码：
from openpyxl import Workbook
from openpyxl.chart import (Reference, Series, BarChart3D)

wb = Workbook()
ws = wb.active

rows = [
    (None, 2013, 2014),
    ('Apples', 5, 4),
    ('Oranges', 6, 2),
    ('Pears', 8, 3)
]

for row in rows:
    ws.append(row)

data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
titles = Reference(ws, min_col=1, min_row=2, max_row=4)
chart = BarChart3D()
chart.title = '3D Bar Chart'
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)

ws.add_chart(chart, 'E5')
wb.save('bar3d.xlsx')



#  5.4 实战训练
# 1.openpyxl 新建Excel
# 程序示例：
# 3.5.2 openpyxl 新建Excel
def fun3_5_2():
    wb = Workbook()

    # 注意：该函数调用工作表的索引(_active_sheet_index)，默认是0。
    # 除非你修改了这个值，否则你使用该函数一直是在对第一张工作表进行操作。
    ws = wb.active

    # 设置sheet名称
    ws.title = 'New Title'
    # 设置sheet颜色
    ws.sheet_properties.tabColor = '1072BA'
    # 保存表格
    wb.save('保存一个新的excel。xlsx')

# 2.openpyxl 打开已存在Excel
# 程序示例：
# 3.5.3 openpyxl 打开已存在Excel
def fun3_5_3():
    wb = load_workbook('xxx.xlsx')
    # 注意：该函数调用工作表的索引(_active_sheet_index)，默认是0。
    # 除非你修改了这个值，否则你使用该函数一直是在对第一张工作表进行操作。
    ws = wb.active

    # 保存表格
    wb.save('copy.xlsx')

# 3.openpyxl 读写Excel
# 程序示例：
# 3.5.4 openpyxl 读写Excel
def fun3_5_4():
    wb = load_workbook('xxx.xlsx')
    # 注意：该函数调用工作表的索引(_active_sheet_index)，默认是0。
    # 除非你修改了这个值，否则你使用该函数一直是在对第一张工作表进行操作。
    ws = wb.active

    # 读取单元格信息
    cellB2_value = ws['B2'].value

    # 写入单元格
    ws['A1'].value = 'OPENPYXL'
    # 保存表格
    wb.save('copy.xlsx')