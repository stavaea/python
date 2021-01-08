#coding:utf-8
from openpyxl import Workbook,load_workbook
import string
from Util.FormatTime import date_time_chinese

class Excel_r_w():
    def __init__(self, excel_path, sheet_name):
        self.excel_path = excel_path
        self.wb = load_workbook(excel_path)
        self.ws = self.wb[sheet_name]
        # 以下为获取多个列的列号组成的列表，不论列是否有数据
        uper_str = string.ascii_uppercase
        self.col_no_list = list(map(lambda x: x, uper_str))#获取基础列号列表，即['A','B'...'Z']基础列号列表
        for iA in 'ABCD':#在基础列号列表后面加上更多的列，即['A'..'Z';'AA'..'AZ';'BA'...'BZ'...]
            for iu in uper_str:
                self.col_no_list.append(iA+iu)

    # 获取sheet中的最大行数
    def get_max_row(self):
        return self.ws.max_row

    # 获取sheet中的最大列数
    def get_max_col(self):
        sol_no = self.ws.max_column#先获得列号
        # print ('filter=', list(filter(lambda x: x==self.col_no_list[sol_no-1], self.col_no_list)))
        return list(filter(lambda x: x == self.col_no_list[sol_no-1], self.col_no_list))[0]#再转成字母

    # 获取sheet的最小（起始）行号
    def get_min_row(self):
        return self.ws.min_row
    # 获取sheet的最小（起始）列号
    def get_min_col(self):
        sol_no = self.ws.min_column#先获取列号
        return list(filter(lambda x: x == self.col_no_list[sol_no-1]), self.col_no_list)[0]#再转成字母

    # 获取sheet的第一个值，index为坐标，如'A1','C5'
    def get_value(self, index):
        return self.ws[index].value

    # 通过坐标写入内容，index为坐标，如'A1','F7'，content为写入的内容
    def write_content(self, index, content):
        self.ws[index] = content
        self.wb.save(self.excel_path)

    #通过坐标写入当前日期
    def write_datetime(self, index):
        self.ws[index] = date_time_chinese()
        self.wb.save(self.excel_path)

if __name__ == '__main__':
    ins = Excel_r_w(r'execl_path')
    print (ins.get_max_row())
    print (ins.get_max_col())
    print (ins.get_min_row())
    print (ins.get_min_col())
    ins.write_content('C11', '你好')
    ins.write_datetime('C12')