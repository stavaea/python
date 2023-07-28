# -*- coding:utf-8 -*-
# ！/usr/bin/env python
# @Time : 2023/7/26 10:47
# @Author : waxberry
# @File : handle_excel.py
# @Software : PyCharm

import openpyxl
from base.base_path import *

class HandleExcel:
    def __init__(self, file_name=None, sheet_name=None)
        '''
        没有传路径时,默认使用 wanadriod接口测试用例.xlsx 文件
        :param file_name:  用例文件
        :param sheet_name: 表单名
        '''
        if file_name:
            self.file_path = os.path.join(testdatas_path. file_name)
            self.sheet_name = sheet_name
        else:
            self.file_path = os.path.join(testdatas_path, 'wanadriod接口测试用例.xlsx')
            self.sheet_name = 'case'
        # 创建工作簿,定位表单
        self.wb = openpyxl.load_workbook(self.file_path)
        self.sheet = self.wb[self.sheet_name]
        # 列总数,行总数
        self.ncols = self.sheet.max_column
        self.nrows = self.sheet.max_row

    def cell_value(self, row=1, column=1):
        '''获取列表中数据，默认取出第一行第一列的值'''
        return self.sheet.cell(row, column).value

    def _get_title(self):
        '''私有函数，返回表头列表'''
        title = []
        for column in range(1, self.ncols+1):
            title.append(self.cell_value(1, column))
        return title

    def get_excel_data(self):
        '''
        :return: 返回字典套列表的方式 [{title_url:value1, title_method:value1}, {title_url:value2, title_method:value2}...]
        '''
        finally_data = []
        for row in range(2, self.nrows+1):
            result_dict = {}
            for column in range(1, self.ncols+1):
                result_dict[self._get_title()[column-1]] = self.cell_value(row, column)
            finally_data.append(result_dict)
        return finally_data

    def get_pytestParametrizeData(self):
        '''
        选用这种参数方式,需要使用数据格式 列表套列表 @pytest.mark.parametrize('', [[], []]), 如 @pytest.mark.parametrize(*get_pytestParametrizeData)
        将 finally_data 中的 title 取出,以字符串形式保存,每个title用逗号(,)隔开
        将 finally_data 中的 value 取出,每行数据保存在一个列表,再集合在一个大列表内
        :return: title, data
        '''
        finally_data = self.get_excel_data()
        data = []
        title = ''
        for i in finally_data:
            value_list = []
            key_list = []
            for key, value in i.items():
                value_list.append(value)
                key_list.append(key)
            title = ','.join(key_list)
            data.append(value_list)
        return title, data

    def rewrite_value(self, new_value, case_id, title):
        '''写入excel，存储使用过的数据（参数化后的数据）'''
        row = self.get_row(case_id)
        column = self.get_column(title)
        self.sheet.cell(row, column).value = new_value
        self.wb.save(self.file_path)

    def get_row(self, case_id):
        '''通过执行的case_id获取当前的行号'''
        for row in range(1, self.nrows+1):
            if self.cell_value(row, 1) == case_id:
                return int(row)

    def get_column(self, title):
        '''通过表头给定字段，获取表头所在列'''
        for column in range(1, self.ncols+1)
            if self.cell_value(1, column) == title:
                return int(column)

if __name__ == '__main__':
    r = HandleExcel()
    print(r.get_excel_data())
